import sys
#from nltk import Nltk
from commonTranslator import CommonTranslator
from GetLiteral import Get_Literal
import time
import openpyxl

#main function
def main():
    # A common translation unit to make things a little easier.
    # Returns the translations as {Google: 'abc', Bing:'abc', Youdao:'abc'}
    translator = CommonTranslator()

    # Languages that we will be examining in our test are:
    # Chinese, Japanese, Korean, French, Russian, Portuguese, Spanish and Swedish
    desired_languages = ['zh-CHS', 'ja', 'ko', 'fr', 'ru', 'pt', 'es', 'sv']

    # Select which lines of the input sentences you wish to use
    input_selection = [516, 518]

    # Name of the output file for the translations
    output_file_name = 'translations.xlsx'

    # Name of the input source file containing the input sentences
    input_file_name = 'csci318.xlsx'

    # Open the new excel sheet
    try:
        output_file = openpyxl.load_workbook(output_file_name)
    except:
        output_file = openpyxl.Workbook()

    sheet_names = output_file.get_sheet_names()
    if 'Google' not in sheet_names:
        output_file.create_sheet('Google')
    if 'Bing' not in sheet_names:
        output_file.create_sheet('Bing')
    if 'Youdao' not in sheet_names:
        output_file.create_sheet('Youdao')

    google_sheet =  output_file['Google']
    bing_sheet = output_file['Bing']
    youdao_sheet = output_file['Youdao']

    # Open the oringinal sentences list
    input_file = openpyxl.load_workbook(filename=input_file_name, read_only=True)
    sentence_list = input_file['English']
    try:
        for line_counter in range(input_selection[0], input_selection[1]):
            sentence = sentence_list.cell(row=line_counter+1, column=1).value
            # For each sentence convert to all desired languages and write to the excel sheets.
            row_entry = line_counter + 1
            # Write the original English sentence
            google_sheet.cell(row = row_entry, column = 1).value = sentence
            bing_sheet.cell(row = row_entry, column = 1).value = sentence
            youdao_sheet.cell(row = row_entry, column = 1).value = sentence
            # Do the translation with each translator and record the result for each language
            for counter, language in enumerate(desired_languages):
                translations = translator.translate('All', sentence, language, 'EN')
                google_sheet.cell(row = row_entry, column = counter + 2).value = translations['Google']
                bing_sheet.cell(row = row_entry, column = counter + 2).value = translations['Bing']
                youdao_sheet.cell(row = row_entry, column = counter + 2).value = translations['Youdao']

    except Exception as exception:
        print (exception)

    finally:
        # Save the file and notify the user
        output_file.save(output_file_name)
        print('Translations have been completed')

    # nltk
    #ownnltk = Nltk()
    #ownnltk.checkScore()

if __name__ == "__main__":
  main()
