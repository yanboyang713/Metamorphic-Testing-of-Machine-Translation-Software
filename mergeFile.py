import openpyxl


def main():
    # used to merge a excel file: "read_file_2" into another excel file: "read_file_1"
    read_file_1=openpyxl.load_workbook(filename='translations.xlsx')
    file1_google_sheet=read_file_1['Google']
    file1_bing_sheet=read_file_1['Bing']
    file1_youdao_sheet=read_file_1['Youdao']


    read_file_2=openpyxl.load_workbook(filename='finial.xlsx')
    file2_google_sheet=read_file_2['Google']
    file2_bing_sheet=read_file_2['Bing']
    file2_youdao_sheet=read_file_2['Youdao']


    #merge google translation
    askMergeGoogleSheet = input("Do you want to merge google sheet? (Y/n) : ")
    if askMergeGoogleSheet=='Y':
        #set the the range by entering the sentence number for start and end.
        for i in range(1, 101):
            #for each same row, if there is no sentence in read_file_1 but exist sentences in read_file_2, then merge that into file_1
            if file1_google_sheet.cell(row=i,column=1).value == None:
                if file2_google_sheet.cell(row=i,column=1).value != None:
                        for j in range(1,10):
                            file1_google_sheet.cell(row=i,column=j).value = file2_google_sheet.cell(row=i,column=j).value


    # merge bing translation
    askMergeBingSheet = input("Do you want to merge bing sheet? (Y/n) : ")
    if askMergeBingSheet=='Y':
        # set the the range by entering the sentence number for start and end.
        for i in range(1, 101):
            # for each same row, if there is no sentence in read_file_1 but exist sentences in read_file_2, then merge that into file_1
            if file1_bing_sheet.cell(row=i,column=1).value == None:
                if file2_bing_sheet.cell(row=i,column=1).value != None:
                        for j in range(1,10):
                            file1_bing_sheet.cell(row=i,column=j).value = file2_bing_sheet.cell(row=i,column=j).value

    # merge youdao translation
    askMergeYoudaoSheet = input("Do you want to merge youdao sheet? (Y/n) : ")
    if askMergeYoudaoSheet=='Y':
        # set the the range by entering the sentence number for start and end.
        for i in range(1, 101):
            # for each same row, if there is no sentence in read_file_1 but exist sentences in read_file_2, then merge that into file_1
            if file1_youdao_sheet.cell(row=i,column=1).value == None:
                if file2_youdao_sheet.cell(row=i,column=1).value != None:
                        for j in range(1,10):
                            file1_youdao_sheet.cell(row=i,column=j).value = file2_youdao_sheet.cell(row=i,column=j).value

    read_file_1.save('translations.xlsx')
    print('Translations have been completed')


if __name__ == "__main__":
  main()

