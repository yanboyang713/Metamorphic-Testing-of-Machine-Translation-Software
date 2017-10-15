import openpyxl

def main():
    read_file = openpyxl.load_workbook(filename='translations.xlsx')
    file_google_sheet = read_file['Google']
    file_bing_sheet = read_file['Bing']
    file_youdao_sheet = read_file['Youdao']


    #validation the google sheet,if only all cells except the first cell are empty then set the whole row empty
    askValidateGoogleSheet = input("Do you want to validate google sheet? (Y/n) : ")
    if askValidateGoogleSheet == 'Y':
        # have already set the size to 1000
        for file_rows in range(1,1001):
            valid_flag=True
            for file_columns in range (2,10):
                if file_google_sheet.cell(row=file_rows,column=file_columns).value is not None:
                    valid_flag=False
            if valid_flag == True:
                for file_columns_2 in range(1, 10):
                    file_google_sheet.cell(row=file_rows, column=file_columns_2).value = None


    #validation the bing sheet,if only all cells except the first cell are empty then set the whole row empty
    askValidateBingSheet = input("Do you want to validate bing sheet? (Y/n) : ")
    if askValidateBingSheet == 'Y':
        #have already set the size to 1000
        for file_rows in range(1,1001):
            valid_flag=True
            for file_columns in range (2,10):
                if file_bing_sheet.cell(row=file_rows,column=file_columns).value is not None:
                    valid_flag=False
            if valid_flag == True:
                for file_columns_2 in range(1, 10):
                    file_bing_sheet.cell(row=file_rows, column=file_columns_2).value = None


    #validation the youdao sheet,if only all cells except the first cell are empty then set the whole row empty
    askValidateYoudaoSheet = input("Do you want to validate youdao sheet? (Y/n) : ")
    if askValidateYoudaoSheet == 'Y':
        # have already set the size to 1000
        for file_rows in range(1,1001):
            valid_flag=True
            for file_columns in range (2,10):
                if file_youdao_sheet.cell(row=file_rows,column=file_columns).value is not None:
                    valid_flag=False
            if valid_flag == True:
                for file_columns_2 in range(1, 10):
                    file_youdao_sheet.cell(row=file_rows, column=file_columns_2).value = None




    read_file.save('translations.xlsx')
    print('Translations have been completed')


if __name__ == "__main__":
  main()