import openpyxl

"""
A standardised class for accessing the translation data.
This class is designed to standardise the reading/writing of translations to the
translation file.

Example Usage:
	testFile = FileManager('test_file.xlsx')
	testFile.writeRow('Google', 1, ["This little piggy..", "went \'wee wee wee\'.."], ['en', 'sv'])
	testFile.writeCell('Youdao', 1, 'all the way home.', 'ko')
	testFile.save()
	sentence = testFile.readCell('Youdao', 300, 'ja')

Generalised for other files, but default initialisation is for reading/writing 'translations.xlsx'
"""
class FileManager:
	def __init__(self, fname = 'translations.xlsx', columns = [['en', 'zh-CHS', 'ja', 'ko', 'fr', 'ru', 'pt', 'es', 'sv']], sheets = ['Google', 'Bing', 'Youdao'], write_header = False):
		self._fname = fname
		if len(columns) == 1:
			# Only one list of columns specified, so assume repeat column names for each sheet.
			self._headers = dict(zip(sheets, [columns][0]*len(sheets)))

		elif len(columns) == len(sheets):
			# Each sheet has a different set of column names.
			self._headers = dict(zip(sheets, columns))

		else:
			raise ValueError('Column length is nonsingular but does not match number of sheets.')

		try:
			self._dataFile = openpyxl.load_workbook(fname)
		except:
			print('(FileManager) \'%s\' does not exist. constructing file..' % (fname))
			self._dataFile = openpyxl.Workbook()
			for sheet in sheets:
				self._dataFile.create_sheet(sheet)
				if write_header:
					self.writeRow(sheet, 1, self._headers[sheet])

	def getColumnNames(self, sheet):
		return self._headers[sheet]

	def getSheetNames(self):
		return self._sheets

	# Returns selected entries from a given Row, or the entire row.
	# Return type is a list whose order matches the requested order of columns.
	def readRow(self, sheet, row_number, columns = None):
		if columns is None:
			columns = self._headers[sheet]

		return [self._dataFile[sheet].cell(row = row_number, column = self._headers[sheet].index(c)+1).value for c in columns]

	# Returns a single entry from a given row.
	# Return type is a string.
	def readCell(self, sheet, row_number, column):
		return self.readRow(sheet, row_number, [column])[0]

	# Writes a set of values to the file for a given row.
	# Can be the entire row of values, or a subset.
	def writeRow(self, sheet, row_number, values, columns = None):
		if columns is None:
			columns = self._headers[sheet]

		if len(values) == len(columns):
			for index, v in enumerate(values):
				self._dataFile[sheet].cell(row = row_number, column = self._headers[sheet].index(columns[index])+1).value = v
		else:
			raise ValueError('Values to write and columns to access not of same length.')

	# Adds a row entry to the last line of the sheet.
	# NOTE: max_row is 1 even when first row is empty, meaning appendEntry() will
	# skip the first row on an enpty spread sheet, but work as normal thereafter.
	# -> should use witeRow(.., 1,..) for first entry or just initialise with header values.
	def appendEntry(self, sheet, values, columns = None):
		self.writeRow(sheet, self._dataFile[sheet].max_row + 1, values, columns)

	def writeCell(self, sheet, row_number, value, column):
		self.writeRow(sheet, row_number, [value], [column])

	def save(self):
		self._dataFile.save(self._fname)
